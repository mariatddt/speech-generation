# Syncs missing translations in TextToSpeechSpreadsheet.xlsx with the translation-toolkit repository.
# Usage: python sync_tts_sheet.py
# Run this before 'dotnet run' when a new phrase or langauge column is added

import re
import argparse
import openpyxl
from pathlib import Path
from translation_toolkit.models import LocalizationDataset
from translation_toolkit.translator import Translator

# config
TOOLS_DIR = Path(__file__).resolve().parent.parent
DEFAULT_KEY_PATH = Path(__file__).resolve().parent / ".cache" / "service-account-key.json"

SHEET_PATH = TOOLS_DIR.parent / "TextToSpeechSpreadsheet.xltx"
SHEET_NAME = "Sheet1"
PHRASE_ID_COLUMN = "AudioFileName"
SOURCE_LOCALE = "en-US"

TEXT_COL_RE = re.compile(r"^(.+)_Text$")
VOICE_COL_RE = re.compile(r"^(.+)_Voice$")

LOCALE_OVERRIDES = {
    "zh-HK": "zh-TW"
}


def normalised_headers(headers: dict) -> dict:
    fixed = dict(headers)
    if "fr-FR-Text" in fixed and "fr-FR_Text" not in fixed:
        fixed["fr-FR_Text"] = fixed.pop("fr-FR-Text")
        print("NOTE: normalized 'fr-FR-Text' -> 'fr-FR_Text'. Fix this typo in the actual sheet.")
    return fixed


def google_code_for(tts_locale: str) -> str:
    return LOCALE_OVERRIDES.get(tts_locale, tts_locale.split("-")[0])


def load_sheet(path: str, sheet_name: str):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook[sheet_name]
    headers = {(cell.value.strip() if isinstance(cell.value, str) else cell.value): cell.column for cell in worksheet[1]}
    return workbook, worksheet, headers


def cell_text(worksheet, row, col):
    value = worksheet.cell(row = row, column = col).value
    return value.strip() if isinstance(value, str) else value

def detect_text_locales(headers: dict) -> list[str]:
    locales = [m.group(1) for h in headers if (m := TEXT_COL_RE.match(h or ""))]
    return [loc for loc in locales if loc != SOURCE_LOCALE]


def build_dataset_from_sheet(worksheet, headers: dict, target_locales: list[str]):
    dataset = LocalizationDataset()
    row_lookup = {}

    source_col = headers.get(f"{SOURCE_LOCALE}_Text")
    if source_col is None:
        raise KeyError(f"Missing source text column '{SOURCE_LOCALE}_Text' in sheet headers: {headers}")

    for row in worksheet.iter_rows(min_row = 2):
        row_idx = row[0].row
        phrase_id = cell_text(worksheet, row_idx, headers[PHRASE_ID_COLUMN])
        if not phrase_id:
            continue
        row_lookup[phrase_id] = row_idx

        source_text = cell_text(worksheet, row_idx, headers[PHRASE_ID_COLUMN])
        if source_text:
            dataset.add_entry("tts_sheet", phrase_id, SOURCE_LOCALE, source_text)

        for locale in target_locales:
            col = headers.get(f"{locale}_Text")
            if col is None:
                continue
            value = worksheet.cell(row = row_idx, column = col).value
            if value:
                dataset.add_entry("tts_sheet", phrase_id, locale, value)

    return dataset, row_lookup


def preview_change(worksheet, headers, dataset, row_lookup, target_locales):
    table = dataset.get_table("tts_sheet")
    if not table:
        return []

    changes = []
    for phrase_id, locales in table.entries.items():
        row_idx = row_lookup.get(phrase_id)
        if row_idx is None:
            continue
        source_text = (locales.get(SOURCE_LOCALE) or "").strip()
        if not source_text:
            continue
        for locale in target_locales:
            if not (locales.get(locale) or "").strip():
                changes.append((phrase_id, locale))

    return changes


def write_back(worksheet, headers, dataset, row_lookup, target_locales):
    table = dataset.get_table("tts_sheet")
    if not table:
        return 0

    updated = 0
    for phrase_id, locales in table.entries.items():
        row_idx = row_lookup.get(phrase_id)
        if row_idx is None:
            continue
        for locale in target_locales:
            text_col = headers.get(f"{locale}_Text")
            if text_col is None:
                continue
            value = locales.get(locale)
            existing = worksheet.cell(row=row_idx, column=text_col).value
            if value and not existing:
                worksheet.cell(row=row_idx, column=text_col, value=value)
                updated += 1

    return updated


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action = "store_true", help = "Print changes without writing the file")
    parser.add_argument(
        "--key",
        default = str(DEFAULT_KEY_PATH),
        help = f"Path to Google service account JSON (default: {DEFAULT_KEY_PATH})",
    )
    args = parser.parse_args()

    workbook, worksheet, headers = load_sheet(SHEET_PATH, SHEET_NAME)
    headers = normalised_headers(headers)

    voice_cols = {m.group(1) for header in headers if (m := VOICE_COL_RE.match(header or ""))}
    text_cols = {m.group(1) for header in headers if (m := TEXT_COL_RE.match(header or ""))}
    assert voice_cols == text_cols, f"Mismatched _Voice/_Text pairs: {voice_cols ^ text_cols}"

    target_locales = detect_text_locales(headers)
    dataset, row_lookup = build_dataset_from_sheet(worksheet, headers, target_locales)

    if args.dry_run:
        changes = preview_change(worksheet, headers, dataset, row_lookup, target_locales)
        print(f"Detected locales: {target_locales}")
        print(f"Would update {len(changes)} cells")
        for phrase_id, locale in changes[:20]:
            print(f" - {phrase_id} [{locale}]")
        return

    key_path = Path(args.key)
    if not key_path.is_file():
        raise FileNotFoundError(
            f"Service account key not found at {key_path}. "
            f"Place it there, or pass --key <path> explicitly."
        )

    google_locale_map = {loc: google_code_for(loc) for loc in target_locales + [SOURCE_LOCALE]}
    translator = Translator(service_account_path = str(key_path), google_locale_map = google_locale_map)
    translator.auto_translate(dataset, target_locales = target_locales, source_locale = SOURCE_LOCALE)

    updated = write_back(worksheet, headers, dataset, row_lookup, target_locales)
    out_path = str(Path(SHEET_PATH).with_name(Path(SHEET_PATH).stem + "_synced.xlsx"))
    
    # Fix Excel corrupt file issue when saving from an .xltx template
    workbook.template = False
    
    workbook.save(out_path)
    print(f"Detected locales: {target_locales}")
    print(f"Updated {updated} cells. Saved to {out_path}")


if __name__ == "__main__":
    main()